"""
CCS Cashflow Assistant — Backend FastAPI
Plugin de Pinokio para gestión de flujos de caja con IA local para PYMEs.

Seguridad aplicada:
  - Sanitización de IDs (prevención de path traversal)
  - CORS restringido a localhost
  - Rate limiting en endpoints de generación
  - Validación de tamaño de inputs
  - Nombres de archivo sanitizados en exportación
  - Datos financieros anonimizados en logs
  - Auto-recuperación de Ollama si no está corriendo
"""

import os
import sys
import json
import uuid
import re
import shutil
import asyncio
import logging
import threading
import argparse
import subprocess
import time
import unicodedata
import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any
from collections import defaultdict

# Forzar UTF-8 en stdout/stderr para Windows
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")

import requests as http_requests
from fastapi import FastAPI, HTTPException, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel, validator

# ---------------------------------------------------------------------------
# Configuración de rutas (siempre absolutas desde __file__)
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent.parent.resolve()
APP_DIR = BASE_DIR / "app"
_raw_data_dir = os.environ.get("DATA_DIR", "")
if _raw_data_dir and "{{" not in _raw_data_dir and Path(_raw_data_dir).is_absolute():
    DATA_DIR = Path(_raw_data_dir)
else:
    DATA_DIR = BASE_DIR / "data"
    if _raw_data_dir and "{{" in _raw_data_dir:
        import logging as _early_log
        _early_log.getLogger(__name__).warning(
            "DATA_DIR contiene plantilla Pinokio sin resolver: %s. "
            "Usando fallback: %s", _raw_data_dir, DATA_DIR
        )
DEFAULTS_DIR = BASE_DIR / "defaults"

def _parse_port():
    """Obtener puerto: 1) argumento --port, 2) env PORT, 3) default 7860."""
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--port", type=int, default=None)
    args, _ = parser.parse_known_args()
    if args.port is not None:
        return args.port
    raw = os.environ.get("PORT", "7860")
    try:
        return int(raw)
    except (ValueError, TypeError):
        return 7860

PORT = _parse_port()
OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://127.0.0.1:11434")

# Timeouts configurables (config.json > env var > default)
def _get_timeout(task_type: str = "default") -> int:
    """Lee timeout desde config.json, luego env var, luego default."""
    config = load_json(DATA_DIR / "config.json", {})
    timeout_map = {
        "default": ("OLLAMA_TIMEOUT", 300),
        "analysis": ("OLLAMA_TIMEOUT_ANALYSIS", 600),
        "simulation": ("OLLAMA_TIMEOUT_SIMULATION", 180),
    }
    env_key, default_val = timeout_map.get(task_type, ("OLLAMA_TIMEOUT", 300))
    config_key = f"ollama_timeout_{task_type}" if task_type != "default" else "ollama_timeout"
    return int(config.get(config_key) or os.getenv(env_key) or default_val)

# Límites de seguridad
MAX_MESSAGE_LENGTH = 10000
MAX_COMPANY_NAME_LENGTH = 200
MAX_MONTHS = 60
RATE_LIMIT_WINDOW = 60
RATE_LIMIT_MAX_REQUESTS = 20

# ---------------------------------------------------------------------------
# Logging (sin datos financieros sensibles)
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("cashflow-assistant")

# ---------------------------------------------------------------------------
# Seguridad: Sanitización de IDs y nombres de archivo
# ---------------------------------------------------------------------------
_SAFE_ID_PATTERN = re.compile(r"^[a-zA-Z0-9_-]{1,64}$")
_SAFE_FILENAME_CHARS = re.compile(r"[^\w\s\-.]", re.UNICODE)

def _sanitize_id(value: str) -> str:
    """Valida que un ID sea seguro (previene path traversal)."""
    if not value or not _SAFE_ID_PATTERN.match(value):
        raise HTTPException(400, f"ID inválido: solo se permiten caracteres alfanuméricos, guiones y guiones bajos (máx 64 chars).")
    return value

def _sanitize_filename(name: str) -> str:
    """Sanitiza un nombre para uso seguro en nombres de archivo."""
    if not name:
        return "empresa"
    name = unicodedata.normalize("NFKD", name)
    name = _SAFE_FILENAME_CHARS.sub("", name)
    name = name.strip().replace(" ", "_")
    name = name[:50] if name else "empresa"
    reserved = {"CON", "PRN", "AUX", "NUL", "COM1", "COM2", "COM3", "COM4",
                "COM5", "COM6", "COM7", "COM8", "COM9", "LPT1", "LPT2",
                "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9"}
    if name.upper().split(".")[0] in reserved:
        name = f"_{name}"
    return name

# ---------------------------------------------------------------------------
# Rate Limiting simple (en memoria)
# ---------------------------------------------------------------------------
_rate_limit_store: Dict[str, list] = defaultdict(list)

def _check_rate_limit(key: str, max_requests: int = RATE_LIMIT_MAX_REQUESTS, window: int = RATE_LIMIT_WINDOW):
    """Verifica rate limit por clave. Lanza HTTPException 429 si se excede."""
    now = time.time()
    _rate_limit_store[key] = [t for t in _rate_limit_store[key] if now - t < window]
    if len(_rate_limit_store[key]) >= max_requests:
        raise HTTPException(429, f"Demasiadas solicitudes. Espera {window} segundos.")
    _rate_limit_store[key].append(now)

# ---------------------------------------------------------------------------
# FastAPI App
# ---------------------------------------------------------------------------
app = FastAPI(title="CCS Cashflow Assistant", version="0.3.0")

# CORS restringido a localhost (Pinokio siempre corre en localhost)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:*",
        "http://127.0.0.1:*",
        "http://0.0.0.0:*",
        "null",
    ],
    allow_origin_regex=r"^https?://(localhost|127\.0\.0\.1|0\.0\.0\.0)(:\d+)?$",
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
    allow_credentials=False,
)

# ---------------------------------------------------------------------------
# Utilidades de persistencia
# ---------------------------------------------------------------------------
def save_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def load_json(path: Path, default=None):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return default if default is not None else {}
    return default if default is not None else {}

# ---------------------------------------------------------------------------
# Utilidades de Ollama con auto-recuperación
# ---------------------------------------------------------------------------
def _fix_encoding(text: str) -> str:
    """Repara texto UTF-8 mal interpretado como latin-1."""
    try:
        return text.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return text

def _extract_json_from_llm(text: str):
    """Extrae JSON robusto de respuestas del LLM."""
    if not text:
        return None
    # Estrategia 1: strip de bloques ```json ... ```
    clean = re.sub(r"```(?:json)?\s*", "", text).strip().rstrip("`").strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass
    # Estrategia 2: parser balanceado
    start = text.find("{")
    if start != -1:
        depth, in_str, escape = 0, False, False
        for i, ch in enumerate(text[start:], start):
            if escape:
                escape = False
                continue
            if ch == "\\" and in_str:
                escape = True
                continue
            if ch == '"':
                in_str = not in_str
                continue
            if not in_str:
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(text[start:i+1])
                        except json.JSONDecodeError:
                            break
    # Estrategia 3: regex greedy
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass
    return None

def ensure_ollama_running() -> bool:
    """Verifica que Ollama esté corriendo; intenta iniciarlo si no lo está."""
    try:
        r = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=3)
        return r.status_code == 200
    except Exception:
        pass
    try:
        logger.info("Ollama no responde. Intentando iniciar...")
        if sys.platform == "win32":
            subprocess.Popen(
                ["ollama", "serve"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NO_WINDOW
            )
        else:
            subprocess.Popen(
                ["ollama", "serve"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL
            )
        for attempt in range(5):
            time.sleep(3)
            try:
                r = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=3)
                if r.status_code == 200:
                    logger.info("Ollama iniciado correctamente.")
                    return True
            except Exception:
                continue
        logger.warning("No se pudo iniciar Ollama automáticamente.")
        return False
    except FileNotFoundError:
        logger.error("Ollama no está instalado en este sistema.")
        return False
    except Exception as e:
        logger.error(f"Error al intentar iniciar Ollama: {e}")
        return False

_pull_status: dict = {}

def _is_model_available(model: str) -> bool:
    try:
        r = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        models = [m["name"] for m in r.json().get("models", [])]
        return model in models or any(m.startswith(model.split(":")[0]) for m in models)
    except Exception:
        return False

def _start_pull_background(model: str):
    if model in _pull_status and _pull_status[model].get("status") in ("pulling", "queued"):
        return
    _pull_status[model] = {"status": "queued", "progress": 0, "error": None}
    threading.Thread(target=_do_pull, args=(model,), daemon=True).start()

def _do_pull(model: str):
    _pull_status[model]["status"] = "pulling"
    try:
        with http_requests.post(f"{OLLAMA_URL}/api/pull",
                                json={"name": model, "stream": True},
                                stream=True, timeout=3600) as r:
            for line in r.iter_lines():
                if line:
                    data = json.loads(line)
                    if "completed" in data and "total" in data and data["total"] > 0:
                        _pull_status[model]["progress"] = int(data["completed"] / data["total"] * 100)
        _pull_status[model] = {"status": "done", "progress": 100, "error": None}
    except Exception as e:
        _pull_status[model] = {"status": "error", "progress": 0, "error": str(e)}

def call_ollama(model: str, system_prompt: str, user_message: str,
                temperature: float = 0.7, timeout: int = None) -> str:
    if timeout is None:
        timeout = _get_timeout("default")
    if not ensure_ollama_running():
        raise HTTPException(503, "Ollama no está disponible. Verifica que esté instalado y ejecutándose.")
    try:
        resp = http_requests.post(
            f"{OLLAMA_URL}/api/chat",
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ],
                "options": {"temperature": temperature},
                "stream": False
            },
            timeout=timeout
        )
        resp.encoding = "utf-8"
        if resp.status_code == 404:
            _start_pull_background(model)
            raise HTTPException(503, f"Modelo {model} no disponible. Descarga iniciada.")
        resp.raise_for_status()
        content = resp.json()["message"]["content"]
        return _fix_encoding(content)
    except http_requests.exceptions.ConnectionError:
        raise HTTPException(503, "No se puede conectar con Ollama. Verifica que esté ejecutándose.")
    except http_requests.exceptions.Timeout:
        raise HTTPException(504, "Timeout al comunicarse con Ollama. Intenta de nuevo.")

def call_ollama_chat(model: str, messages: list, temperature: float = 0.7, timeout: int = None) -> str:
    """Llamada a Ollama con historial de mensajes completo."""
    if timeout is None:
        timeout = _get_timeout("default")
    if not ensure_ollama_running():
        raise HTTPException(503, "Ollama no está disponible.")
    try:
        resp = http_requests.post(
            f"{OLLAMA_URL}/api/chat",
            json={
                "model": model,
                "messages": messages,
                "options": {"temperature": temperature},
                "stream": False
            },
            timeout=timeout
        )
        resp.encoding = "utf-8"
        if resp.status_code == 404:
            _start_pull_background(model)
            raise HTTPException(503, f"Modelo {model} no disponible. Descarga iniciada.")
        resp.raise_for_status()
        content = resp.json()["message"]["content"]
        return _fix_encoding(content)
    except http_requests.exceptions.ConnectionError:
        raise HTTPException(503, "No se puede conectar con Ollama.")
    except http_requests.exceptions.Timeout:
        raise HTTPException(504, "Timeout al comunicarse con Ollama.")

# ---------------------------------------------------------------------------
# Carga de agentes y prompts
# ---------------------------------------------------------------------------
def get_agents() -> dict:
    return load_json(DATA_DIR / "agents" / "agents.json", {"version": "0.2.0", "agents": []})

def get_agent(agent_id: str) -> dict:
    _sanitize_id(agent_id)
    agents = get_agents()
    for a in agents.get("agents", []):
        if a["id"] == agent_id:
            return a
    return None

def get_prompt(prompt_name: str) -> str:
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "", prompt_name)
    if not safe_name:
        return ""
    path = DATA_DIR / "prompts" / "system" / f"{safe_name}.md"
    if path.exists():
        return path.read_text(encoding="utf-8")
    path = DEFAULTS_DIR / "prompts" / f"{safe_name}.md"
    if path.exists():
        return path.read_text(encoding="utf-8")
    return ""

# ---------------------------------------------------------------------------
# Modelos Pydantic con validación
# ---------------------------------------------------------------------------
class CompanyCreate(BaseModel):
    name: str
    sector: str = ""
    description: str = ""

    @validator("name")
    def validate_name(cls, v):
        if not v or not v.strip():
            raise ValueError("El nombre de la empresa es requerido.")
        if len(v) > MAX_COMPANY_NAME_LENGTH:
            raise ValueError(f"El nombre no puede exceder {MAX_COMPANY_NAME_LENGTH} caracteres.")
        return v.strip()

class ChatMessage(BaseModel):
    company_id: str
    message: str
    session_id: Optional[str] = ""

    @validator("session_id", pre=True, always=True)
    def coerce_session_id(cls, v):
        if v is None:
            return ""
        return str(v)

    @validator("message")
    def validate_message(cls, v):
        if not v or not v.strip():
            raise ValueError("El mensaje no puede estar vacío.")
        if len(v) > MAX_MESSAGE_LENGTH:
            raise ValueError(f"El mensaje no puede exceder {MAX_MESSAGE_LENGTH} caracteres.")
        return v.strip()

class ScenarioRequest(BaseModel):
    instruction: str
    use_ai: Optional[bool] = False
    params: Optional[Dict[str, Any]] = None

    @validator("instruction")
    def validate_instruction(cls, v):
        if not v or not v.strip():
            raise ValueError("La instrucción no puede estar vacía.")
        if len(v) > MAX_MESSAGE_LENGTH:
            raise ValueError(f"La instrucción no puede exceder {MAX_MESSAGE_LENGTH} caracteres.")
        return v.strip()

class AgentConfigUpdate(BaseModel):
    model: Optional[str] = None
    temperature: Optional[float] = None
    system_prompt: Optional[str] = None

    @validator("temperature")
    def validate_temperature(cls, v):
        if v is not None and (v < 0.0 or v > 2.0):
            raise ValueError("La temperatura debe estar entre 0.0 y 2.0.")
        return v

    @validator("system_prompt")
    def validate_system_prompt(cls, v):
        if v is not None and len(v) > 50000:
            raise ValueError("El prompt del sistema no puede exceder 50000 caracteres.")
        return v

class MonthData(BaseModel):
    month: Optional[str] = None
    label: Optional[str] = None
    income: Optional[dict] = None
    expenses: Optional[dict] = None

# ---------------------------------------------------------------------------
# Normalización de datos de flujo de caja
# ---------------------------------------------------------------------------
def _to_num(val) -> float:
    """Convierte un valor a número de forma robusta."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        clean = val.strip().replace("$", "").replace(" ", "")
        if clean.count(".") > 1:
            clean = clean.replace(".", "")
        elif "." in clean and "," in clean:
            clean = clean.replace(".", "").replace(",", ".")
        elif "," in clean and "." not in clean:
            parts = clean.split(",")
            if all(len(p) == 3 for p in parts[1:]):
                clean = clean.replace(",", "")
            else:
                clean = clean.replace(",", ".")
        clean = re.sub(r"[^\d.\-]", "", clean)
        try:
            return float(clean) if clean else 0.0
        except ValueError:
            return 0.0
    return 0.0

def _normalize_month(m: dict) -> dict:
    """Normaliza un mes individual: asegura estructura y recalcula totales."""
    income = m.get("income", {})
    if not isinstance(income, dict):
        income = {}
    expenses = m.get("expenses", {})
    if not isinstance(expenses, dict):
        expenses = {}

    sales = _to_num(income.get("sales", 0))
    other_income = _to_num(income.get("other_income", 0))
    income_total = sales + other_income

    variable_costs = _to_num(expenses.get("variable_costs", 0))
    fixed_costs = _to_num(expenses.get("fixed_costs", 0))
    variable_expenses = _to_num(expenses.get("variable_expenses", 0))
    debt_payments = _to_num(expenses.get("debt_payments", 0))
    taxes = _to_num(expenses.get("taxes", 0))
    investments = _to_num(expenses.get("investments", 0))
    expenses_total = variable_costs + fixed_costs + variable_expenses + debt_payments + taxes + investments

    net_flow = income_total - expenses_total

    return {
        "month": m.get("month", ""),
        "label": m.get("label", m.get("month", "")),
        "income": {
            "sales": sales,
            "other_income": other_income,
            "total": income_total
        },
        "expenses": {
            "variable_costs": variable_costs,
            "fixed_costs": fixed_costs,
            "variable_expenses": variable_expenses,
            "debt_payments": debt_payments,
            "taxes": taxes,
            "investments": investments,
            "total": expenses_total
        },
        "net_flow": net_flow,
        "cumulative_balance": 0
    }

def _rescue_cashflow_structure(data: dict) -> dict:
    """Busca la estructura de meses en cualquier nivel del JSON del LLM.
    Los LLMs a veces envuelven el resultado en campos extra como 'data', 'cashflow',
    'flujo_de_caja', 'result', etc. Esta función busca recursivamente."""
    if not isinstance(data, dict):
        return data
    
    # Si ya tiene months como lista con al menos un elemento, está bien
    months = data.get("months", None)
    if isinstance(months, list) and len(months) > 0:
        # Verificar que los elementos parecen meses (tienen income o expenses)
        first = months[0] if months else {}
        if isinstance(first, dict) and ("income" in first or "expenses" in first or "sales" in first or "ingresos" in first):
            return data
    
    # Buscar en campos comunes donde el LLM podría haber anidado los datos
    candidate_keys = ["data", "cashflow", "flujo_de_caja", "result", "resultado",
                      "flujo", "cash_flow", "projection", "proyeccion"]
    for key in candidate_keys:
        if key in data and isinstance(data[key], dict):
            inner = data[key]
            if isinstance(inner.get("months"), list) and len(inner["months"]) > 0:
                # Mover campos del wrapper al nivel superior
                for k, v in inner.items():
                    data[k] = v
                return data
    
    # Buscar cualquier campo que sea una lista de dicts con estructura de mes
    for key, val in data.items():
        if key == "months":
            continue
        if isinstance(val, list) and len(val) > 0:
            first = val[0] if val else {}
            if isinstance(first, dict) and ("income" in first or "expenses" in first or
                                            "sales" in first or "ingresos" in first or
                                            "month" in first or "mes" in first):
                data["months"] = val
                return data
    
    # Buscar meses con nombres en español
    if "meses" in data and isinstance(data["meses"], list):
        data["months"] = data["meses"]
        return data
    
    return data


def _normalize_month_from_spanish(m: dict) -> dict:
    """Intenta normalizar un mes que puede venir con campos en español."""
    if not isinstance(m, dict):
        return m
    
    # Mapear campos en español a inglés
    if "ingresos" in m and "income" not in m:
        inc = m["ingresos"]
        if isinstance(inc, dict):
            m["income"] = {
                "sales": _to_num(inc.get("ventas", inc.get("sales", 0))),
                "other_income": _to_num(inc.get("otros_ingresos", inc.get("otros", inc.get("other_income", 0)))),
            }
        elif isinstance(inc, (int, float, str)):
            m["income"] = {"sales": _to_num(inc), "other_income": 0}
    
    if "gastos" in m and "expenses" not in m:
        exp = m["gastos"]
        if isinstance(exp, dict):
            m["expenses"] = {
                "variable_costs": _to_num(exp.get("costos_variables", exp.get("variable_costs", 0))),
                "fixed_costs": _to_num(exp.get("costos_fijos", exp.get("fixed_costs", 0))),
                "variable_expenses": _to_num(exp.get("gastos_variables", exp.get("variable_expenses", 0))),
                "debt_payments": _to_num(exp.get("deudas", exp.get("debt_payments", exp.get("pagos_deuda", 0)))),
                "taxes": _to_num(exp.get("impuestos", exp.get("taxes", 0))),
                "investments": _to_num(exp.get("inversiones", exp.get("investments", 0))),
            }
        elif isinstance(exp, (int, float, str)):
            m["expenses"] = {"variable_costs": _to_num(exp), "fixed_costs": 0, "variable_expenses": 0, "debt_payments": 0, "taxes": 0, "investments": 0}
    
    if "mes" in m and "month" not in m:
        m["month"] = m["mes"]
    if "etiqueta" in m and "label" not in m:
        m["label"] = m["etiqueta"]
    
    # Si tiene ventas a nivel raíz (sin income wrapper)
    if "sales" in m and "income" not in m:
        m["income"] = {"sales": _to_num(m["sales"]), "other_income": _to_num(m.get("other_income", 0))}
    if "ventas" in m and "income" not in m:
        m["income"] = {"sales": _to_num(m["ventas"]), "other_income": _to_num(m.get("otros_ingresos", 0))}
    
    return m


def normalize_cashflow(data: dict) -> dict:
    """Normaliza un flujo de caja completo: recalcula totales, saldos acumulados y summary."""
    # Primero intentar rescatar la estructura si el LLM la anidó
    data = _rescue_cashflow_structure(data)
    
    months = data.get("months", [])
    if not isinstance(months, list):
        months = []
    months = months[:MAX_MONTHS]
    
    # Normalizar meses que pueden venir en español
    months = [_normalize_month_from_spanish(m) for m in months if isinstance(m, dict)]

    normalized_months = []
    cumulative = 0.0
    total_income = 0.0
    total_expenses = 0.0

    for m in months:
        nm = _normalize_month(m)
        cumulative += nm["net_flow"]
        nm["cumulative_balance"] = cumulative
        total_income += nm["income"]["total"]
        total_expenses += nm["expenses"]["total"]
        normalized_months.append(nm)

    num_months = len(normalized_months)
    net_cashflow = total_income - total_expenses
    avg_balance = cumulative / num_months if num_months > 0 else 0

    data["months"] = normalized_months
    data["summary"] = {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "net_cashflow": net_cashflow,
        "average_monthly_balance": avg_balance,
        "num_months": num_months
    }

    return data

# ---------------------------------------------------------------------------
# Motor de Simulación Local (sin LLM - rápido y determinístico)
# ---------------------------------------------------------------------------
def simulate_local(cashflow: dict, params: dict, task_id: str = None) -> dict:
    """
    Ejecuta simulación matemática local sin necesidad de LLM.
    Parámetros soportados:
      - sales_change_pct: % de cambio en ventas
      - costs_change_pct: % de cambio en costos variables
      - fixed_costs_change_pct: % de cambio en costos fijos
      - inflation_annual_pct: inflación anual (aplicada mensualmente)
      - new_hires: número de nuevas contrataciones
      - hire_cost: costo mensual por contratación (default 800000)
      - tax_change_pct: % de cambio en impuestos
      - debt_change_pct: % de cambio en pagos de deuda
      - investment_change_pct: % de cambio en inversiones
      - other_income_change_pct: % de cambio en otros ingresos
      - start_month: mes desde el cual aplicar cambios (0-based, default 0)
    """
    result = copy.deepcopy(cashflow)
    months = result.get("months", [])
    if not months:
        return result

    # Actualizar progreso
    if task_id:
        _generation_status[task_id]["progress"] = 20
        _generation_status[task_id]["step"] = "Preparando parámetros de simulación..."

    sales_pct = params.get("sales_change_pct", 0) / 100.0
    costs_pct = params.get("costs_change_pct", 0) / 100.0
    fixed_costs_pct = params.get("fixed_costs_change_pct", 0) / 100.0
    inflation_annual = params.get("inflation_annual_pct", 0) / 100.0
    new_hires = params.get("new_hires", 0)
    hire_cost = params.get("hire_cost", 800000)
    tax_pct = params.get("tax_change_pct", 0) / 100.0
    debt_pct = params.get("debt_change_pct", 0) / 100.0
    investment_pct = params.get("investment_change_pct", 0) / 100.0
    other_income_pct = params.get("other_income_change_pct", 0) / 100.0
    start_month = params.get("start_month", 0)

    # Inflación mensual compuesta
    monthly_inflation = (1 + inflation_annual) ** (1/12) - 1 if inflation_annual > 0 else 0

    total_months = len(months)
    changes_applied = []

    if task_id:
        _generation_status[task_id]["progress"] = 30
        _generation_status[task_id]["step"] = "Aplicando cambios a los meses..."

    # Guardar checkpoint inicial
    if task_id:
        _save_simulation_checkpoint(task_id, result, 0, total_months)

    for i in range(total_months):
        if i < start_month:
            continue

        month = months[i]
        income = month.get("income", {})
        expenses = month.get("expenses", {})

        # Calcular factor de inflación acumulada desde start_month
        months_elapsed = i - start_month
        inflation_factor = (1 + monthly_inflation) ** months_elapsed if monthly_inflation > 0 else 1.0

        # Aplicar cambios a ingresos
        if sales_pct != 0:
            income["sales"] = round(income.get("sales", 0) * (1 + sales_pct), 2)
        if other_income_pct != 0:
            income["other_income"] = round(income.get("other_income", 0) * (1 + other_income_pct), 2)

        # Aplicar cambios a gastos
        if costs_pct != 0:
            expenses["variable_costs"] = round(expenses.get("variable_costs", 0) * (1 + costs_pct), 2)
        if fixed_costs_pct != 0:
            expenses["fixed_costs"] = round(expenses.get("fixed_costs", 0) * (1 + fixed_costs_pct), 2)
        if tax_pct != 0:
            expenses["taxes"] = round(expenses.get("taxes", 0) * (1 + tax_pct), 2)
        if debt_pct != 0:
            expenses["debt_payments"] = round(expenses.get("debt_payments", 0) * (1 + debt_pct), 2)
        if investment_pct != 0:
            expenses["investments"] = round(expenses.get("investments", 0) * (1 + investment_pct), 2)

        # Nuevas contrataciones (afecta costos fijos)
        if new_hires > 0:
            expenses["fixed_costs"] = round(expenses.get("fixed_costs", 0) + (new_hires * hire_cost), 2)

        # Inflación (afecta costos fijos y gastos variables progresivamente)
        if monthly_inflation > 0:
            expenses["fixed_costs"] = round(expenses.get("fixed_costs", 0) * inflation_factor, 2)
            expenses["variable_expenses"] = round(expenses.get("variable_expenses", 0) * inflation_factor, 2)

        month["income"] = income
        month["expenses"] = expenses
        months[i] = month

        # Actualizar progreso cada 2 meses
        if task_id and (i % 2 == 0 or i == total_months - 1):
            progress = 30 + int((i / total_months) * 50)
            _generation_status[task_id]["progress"] = progress
            _generation_status[task_id]["step"] = f"Procesando mes {i + 1} de {total_months}..."
            # Guardar checkpoint cada 4 meses
            if i % 4 == 0:
                _save_simulation_checkpoint(task_id, result, i, total_months)

    result["months"] = months

    if task_id:
        _generation_status[task_id]["progress"] = 85
        _generation_status[task_id]["step"] = "Recalculando totales y alertas..."

    # Re-normalizar
    result = normalize_cashflow(result)

    # Generar descripción de cambios
    if sales_pct != 0:
        changes_applied.append(f"Ventas {'aumentadas' if sales_pct > 0 else 'reducidas'} en {abs(int(sales_pct*100))}%")
    if costs_pct != 0:
        changes_applied.append(f"Costos variables {'aumentados' if costs_pct > 0 else 'reducidos'} en {abs(int(costs_pct*100))}%")
    if fixed_costs_pct != 0:
        changes_applied.append(f"Costos fijos {'aumentados' if fixed_costs_pct > 0 else 'reducidos'} en {abs(int(fixed_costs_pct*100))}%")
    if inflation_annual > 0:
        changes_applied.append(f"Inflación anual del {int(inflation_annual*100)}% aplicada")
    if new_hires > 0:
        changes_applied.append(f"{new_hires} nuevas contrataciones (${hire_cost:,.0f} c/u)")
    if tax_pct != 0:
        changes_applied.append(f"Impuestos {'aumentados' if tax_pct > 0 else 'reducidos'} en {abs(int(tax_pct*100))}%")
    if debt_pct != 0:
        changes_applied.append(f"Pagos de deuda {'aumentados' if debt_pct > 0 else 'reducidos'} en {abs(int(debt_pct*100))}%")
    if investment_pct != 0:
        changes_applied.append(f"Inversiones {'aumentadas' if investment_pct > 0 else 'reducidas'} en {abs(int(investment_pct*100))}%")
    if other_income_pct != 0:
        changes_applied.append(f"Otros ingresos {'aumentados' if other_income_pct > 0 else 'reducidos'} en {abs(int(other_income_pct*100))}%")

    # Generar alertas automáticas
    alerts = []
    for i, m in enumerate(result["months"]):
        if m["cumulative_balance"] < 0:
            alerts.append({
                "month": m.get("label", m.get("month", f"Mes {i+1}")),
                "type": "danger",
                "message": f"Saldo acumulado negativo: ${m['cumulative_balance']:,.0f}"
            })
        elif m["net_flow"] < 0:
            alerts.append({
                "month": m.get("label", m.get("month", f"Mes {i+1}")),
                "type": "warning",
                "message": f"Flujo neto negativo: ${m['net_flow']:,.0f}"
            })

    # Generar nombre del escenario
    scenario_name = "Simulación: " + ", ".join(changes_applied[:3])
    if len(changes_applied) > 3:
        scenario_name += f" (+{len(changes_applied)-3} más)"

    # Calcular impacto
    base_summary = cashflow.get("summary", {})
    new_summary = result.get("summary", {})
    impact_income = new_summary.get("total_income", 0) - base_summary.get("total_income", 0)
    impact_expenses = new_summary.get("total_expenses", 0) - base_summary.get("total_expenses", 0)
    impact_net = new_summary.get("net_cashflow", 0) - base_summary.get("net_cashflow", 0)

    result["scenario_name"] = scenario_name
    result["changes_applied"] = changes_applied
    result["alerts"] = alerts
    result["impact_summary"] = {
        "income_change": impact_income,
        "expenses_change": impact_expenses,
        "net_change": impact_net,
        "description": f"Impacto neto: ${impact_net:+,.0f} en el período"
    }
    result["recommendations"] = _generate_recommendations(result, cashflow)

    if task_id:
        _generation_status[task_id]["progress"] = 95
        _generation_status[task_id]["step"] = "Finalizando simulación..."

    return result


def _generate_recommendations(simulated: dict, original: dict) -> list:
    """Genera recomendaciones basadas en la comparación entre simulado y original."""
    recs = []
    sim_summary = simulated.get("summary", {})
    orig_summary = original.get("summary", {})

    net_change = sim_summary.get("net_cashflow", 0) - orig_summary.get("net_cashflow", 0)

    if net_change < 0:
        recs.append("El escenario simulado reduce el flujo neto. Considere medidas compensatorias.")
    if net_change > 0:
        recs.append("El escenario simulado mejora el flujo neto. Evalúe su viabilidad de implementación.")

    # Verificar meses con saldo negativo
    negative_months = [m for m in simulated.get("months", []) if m.get("cumulative_balance", 0) < 0]
    if negative_months:
        recs.append(f"Hay {len(negative_months)} mes(es) con saldo acumulado negativo. Considere líneas de crédito o reducción de gastos.")

    # Verificar tendencia
    months = simulated.get("months", [])
    if len(months) >= 3:
        last_3_flows = [m.get("net_flow", 0) for m in months[-3:]]
        if all(f < 0 for f in last_3_flows):
            recs.append("Los últimos 3 meses muestran flujo negativo sostenido. Revise la estrategia financiera.")
        elif all(f > 0 for f in last_3_flows):
            recs.append("Los últimos 3 meses muestran flujo positivo sostenido. Buen momento para inversiones.")

    return recs


def _save_simulation_checkpoint(task_id: str, data: dict, month_index: int, total_months: int):
    """Guarda un checkpoint de la simulación en disco para recuperación."""
    checkpoints_dir = DATA_DIR / "checkpoints"
    checkpoints_dir.mkdir(parents=True, exist_ok=True)
    checkpoint = {
        "task_id": task_id,
        "timestamp": datetime.now().isoformat(),
        "month_index": month_index,
        "total_months": total_months,
        "partial_data": data
    }
    save_json(checkpoints_dir / f"{task_id}.json", checkpoint)


# ---------------------------------------------------------------------------
# Estado de generación (en memoria)
# ---------------------------------------------------------------------------
_generation_status: dict = {}

# ---------------------------------------------------------------------------
# Endpoints: Health
# ---------------------------------------------------------------------------
@app.get("/api/health")
async def health():
    ollama_ok = False
    ollama_models = []
    try:
        r = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=3)
        if r.status_code == 200:
            ollama_ok = True
            ollama_models = [m["name"] for m in r.json().get("models", [])]
    except Exception:
        pass
    return {
        "status": "ok",
        "ollama": "connected" if ollama_ok else "disconnected",
        "ollama_models": ollama_models,
        "version": "0.3.0",
        "platform": sys.platform
    }

# ---------------------------------------------------------------------------
# Endpoints: Readiness (semáforo de sistema)
# ---------------------------------------------------------------------------
@app.get("/api/readiness")
async def check_readiness():
    issues = []
    ready = True

    ollama_ok = False
    models = []
    try:
        resp = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        if resp.status_code == 200:
            ollama_ok = True
            models = [m["name"] for m in resp.json().get("models", [])]
    except Exception:
        pass

    if not ollama_ok:
        ready = False
        issues.append({
            "type": "ollama_unavailable",
            "message": "Ollama no está disponible. Asegúrate de que esté instalado y corriendo.",
            "severity": "critical",
        })

    if ollama_ok and not models:
        ready = False
        issues.append({
            "type": "no_models",
            "message": "No hay modelos de IA descargados. Se están descargando automáticamente...",
            "severity": "warning",
        })

    active_pulls = []
    for model_name, info in _pull_status.items():
        if info.get("status") in ("queued", "pulling"):
            active_pulls.append({
                "model": model_name,
                "status": info.get("status"),
                "progress": info.get("progress", 0),
            })

    if active_pulls:
        issues.append({
            "type": "models_downloading",
            "message": f"Descargando {len(active_pulls)} modelo(s) de IA...",
            "severity": "info",
            "pulls": active_pulls,
        })

    return {
        "ready": ready and not active_pulls,
        "ollama_available": ollama_ok,
        "models_count": len(models),
        "models": models,
        "active_pulls": active_pulls,
        "issues": issues,
    }

# ---------------------------------------------------------------------------
# Endpoints: Ollama Status
# ---------------------------------------------------------------------------
@app.get("/api/ollama/status")
async def ollama_status():
    try:
        resp = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        models = [m["name"] for m in resp.json().get("models", [])]
        return {"available": True, "models": models}
    except Exception:
        return {"available": False, "models": []}

# ---------------------------------------------------------------------------
# Endpoints: Hardware Performance (semáforo de modelos)
# ---------------------------------------------------------------------------
@app.get("/api/hardware/performance")
async def get_hardware_performance():
    import platform as plat

    hw = {
        "platform": sys.platform,
        "platform_name": plat.system(),
        "architecture": plat.machine(),
        "cpu_count": os.cpu_count() or 1,
        "ram_gb": 0,
        "gpu_name": "No detectada",
        "vram_gb": 0,
    }

    try:
        if sys.platform == "win32":
            import ctypes
            kernel32 = ctypes.windll.kernel32
            c_ulonglong = ctypes.c_ulonglong
            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [("dwLength", ctypes.c_ulong),
                            ("dwMemoryLoad", ctypes.c_ulong),
                            ("ullTotalPhys", c_ulonglong),
                            ("ullAvailPhys", c_ulonglong),
                            ("ullTotalPageFile", c_ulonglong),
                            ("ullAvailPageFile", c_ulonglong),
                            ("ullTotalVirtual", c_ulonglong),
                            ("ullAvailVirtual", c_ulonglong),
                            ("ullAvailExtendedVirtual", c_ulonglong)]
            stat = MEMORYSTATUSEX()
            stat.dwLength = ctypes.sizeof(stat)
            kernel32.GlobalMemoryStatusEx(ctypes.byref(stat))
            hw["ram_gb"] = round(stat.ullTotalPhys / (1024**3), 1)
        elif sys.platform == "darwin":
            out = subprocess.check_output(["sysctl", "-n", "hw.memsize"], timeout=5)
            hw["ram_gb"] = round(int(out.strip()) / (1024**3), 1)
        else:
            with open("/proc/meminfo") as f:
                for line in f:
                    if line.startswith("MemTotal"):
                        kb = int(line.split()[1])
                        hw["ram_gb"] = round(kb / (1024**2), 1)
                        break
    except Exception as e:
        logger.debug(f"No se pudo detectar RAM: {e}")

    try:
        if sys.platform == "darwin" and plat.machine() == "arm64":
            hw["gpu_name"] = f"Apple Silicon ({plat.processor() or 'M-series'})"
            hw["vram_gb"] = hw["ram_gb"]
        else:
            out = subprocess.check_output(
                ["nvidia-smi", "--query-gpu=name,memory.total", "--format=csv,noheader,nounits"],
                timeout=5, stderr=subprocess.DEVNULL
            ).decode().strip()
            if out:
                parts = out.split(",")
                hw["gpu_name"] = parts[0].strip()
                hw["vram_gb"] = round(int(parts[1].strip()) / 1024, 1) if len(parts) > 1 else 0
    except Exception:
        pass

    models_perf = []
    try:
        resp = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        if resp.status_code == 200:
            for m in resp.json().get("models", []):
                model_name = m.get("name", "")
                model_size_bytes = m.get("size", 0)
                model_size_gb = round(model_size_bytes / (1024**3), 1)

                params_b = _estimate_model_params(model_name)
                tps = _estimate_tokens_per_second(
                    params_b, hw["ram_gb"], hw["vram_gb"],
                    hw["cpu_count"], hw["gpu_name"]
                )
                grade = _compute_grade(tps, model_size_gb, hw["ram_gb"])

                models_perf.append({
                    "model": model_name,
                    "size_gb": model_size_gb,
                    "params_b": params_b,
                    "estimated_tps": tps,
                    "grade": grade["grade"],
                    "grade_label": grade["label"],
                    "grade_color": grade["color"],
                    "score": grade["score"],
                    "ram_pct": round((model_size_gb / hw["ram_gb"]) * 100, 1) if hw["ram_gb"] > 0 else 100,
                })
    except Exception as e:
        logger.debug(f"No se pudieron obtener modelos para performance: {e}")

    return {"hardware": hw, "models": models_perf}


def _estimate_model_params(model_name: str) -> float:
    name_lower = model_name.lower()
    match = re.search(r'(\d+\.?\d*)b', name_lower)
    if match:
        return float(match.group(1))
    if "1b" in name_lower or "1.1b" in name_lower:
        return 1.0
    if "3b" in name_lower:
        return 3.0
    if "7b" in name_lower or "8b" in name_lower:
        return 8.0
    if "13b" in name_lower or "14b" in name_lower:
        return 14.0
    if "32b" in name_lower or "34b" in name_lower:
        return 32.0
    if "70b" in name_lower:
        return 70.0
    return 7.0


def _estimate_tokens_per_second(
    params_b: float, ram_gb: float, vram_gb: float,
    cpu_count: int, gpu_name: str
) -> int:
    model_gb = params_b * 0.6
    is_apple_silicon = "apple" in gpu_name.lower() or "m1" in gpu_name.lower() or "m2" in gpu_name.lower() or "m3" in gpu_name.lower() or "m4" in gpu_name.lower()
    has_nvidia = "nvidia" in gpu_name.lower() or "geforce" in gpu_name.lower() or "rtx" in gpu_name.lower()

    if is_apple_silicon:
        available = ram_gb
        if model_gb > available * 0.8:
            return max(1, int(5 * (available / model_gb)))
        bandwidth_factor = min(2.0, ram_gb / 16.0)
        base_tps = 60 / (params_b / 8.0)
        return max(1, int(base_tps * bandwidth_factor))
    elif has_nvidia and vram_gb > 0:
        if model_gb <= vram_gb * 0.9:
            base_tps = 80 / (params_b / 8.0)
            return max(1, int(base_tps))
        elif model_gb <= vram_gb + ram_gb * 0.5:
            return max(1, int(20 / (params_b / 8.0)))
        else:
            return max(1, int(5 / (params_b / 8.0)))
    else:
        if model_gb > ram_gb * 0.7:
            return max(1, int(2 * (ram_gb / model_gb)))
        core_factor = min(2.0, cpu_count / 8.0)
        base_tps = 25 / (params_b / 8.0)
        return max(1, int(base_tps * core_factor))


def _compute_grade(tps: int, model_size_gb: float, ram_gb: float) -> dict:
    if ram_gb > 0 and model_size_gb > ram_gb * 0.9:
        return {"grade": "F", "label": "NO EJECUTABLE", "color": "#dc2626", "score": 0}
    if tps >= 30:
        score = min(100, 80 + int((tps - 30) * 0.5))
        return {"grade": "S", "label": "EXCELENTE", "color": "#22c55e", "score": score}
    elif tps >= 15:
        score = 65 + int((tps - 15) * 1.0)
        return {"grade": "A", "label": "MUY BUENO", "color": "#4ade80", "score": score}
    elif tps >= 8:
        score = 50 + int((tps - 8) * 2.0)
        return {"grade": "B", "label": "ACEPTABLE", "color": "#facc15", "score": score}
    elif tps >= 4:
        score = 30 + int((tps - 4) * 5.0)
        return {"grade": "C", "label": "AJUSTADO", "color": "#f97316", "score": score}
    elif tps >= 2:
        score = 15 + int((tps - 2) * 7.5)
        return {"grade": "D", "label": "MUY LENTO", "color": "#ef4444", "score": score}
    else:
        return {"grade": "F", "label": "NO RECOMENDADO", "color": "#dc2626", "score": max(0, tps * 7)}


# ---------------------------------------------------------------------------
# Endpoints: Empresas (con sanitización de IDs)
# ---------------------------------------------------------------------------
@app.post("/api/companies", status_code=201)
async def create_company(data: CompanyCreate):
    company_id = str(uuid.uuid4())[:8]
    company_dir = DATA_DIR / "companies" / company_id
    company_dir.mkdir(parents=True, exist_ok=True)
    company = {
        "id": company_id,
        "name": data.name,
        "sector": data.sector,
        "description": data.description,
        "status": "pending",
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    save_json(company_dir / "company.json", company)
    logger.info(f"Empresa creada: {company_id}")
    return company

@app.get("/api/companies")
async def list_companies():
    companies_dir = DATA_DIR / "companies"
    companies_dir.mkdir(parents=True, exist_ok=True)
    companies = []
    for d in sorted(companies_dir.iterdir()):
        if d.is_dir():
            company = load_json(d / "company.json")
            if company:
                companies.append(company)
    return {"companies": companies}

@app.get("/api/companies/{company_id}")
async def get_company(company_id: str):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    if not company_dir.exists():
        raise HTTPException(404, "Empresa no encontrada")
    return load_json(company_dir / "company.json")

@app.delete("/api/companies/{company_id}")
async def delete_company(company_id: str):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    if company_dir.exists():
        shutil.rmtree(str(company_dir))
    # Limpiar sesiones
    sessions_dir = DATA_DIR / "sessions"
    if sessions_dir.exists():
        for f in sessions_dir.glob(f"{company_id}_*"):
            f.unlink()
    # Limpiar escenarios
    scenarios_dir = DATA_DIR / "scenarios"
    if scenarios_dir.exists():
        for f in scenarios_dir.glob(f"{company_id}_*"):
            f.unlink()
    return {"status": "deleted"}

# ---------------------------------------------------------------------------
# Endpoints: Chat / Entrevista
# ---------------------------------------------------------------------------
@app.post("/api/chat/interview")
async def chat_interview(data: ChatMessage):
    company_id = _sanitize_id(data.company_id)
    company_dir = DATA_DIR / "companies" / company_id
    if not company_dir.exists():
        raise HTTPException(404, "Empresa no encontrada")

    _check_rate_limit(f"chat_{company_id}")

    company = load_json(company_dir / "company.json")

    session_id = data.session_id or str(uuid.uuid4())[:8]
    session_id = re.sub(r"[^a-zA-Z0-9_-]", "", session_id)[:64]
    session_path = DATA_DIR / "sessions" / f"{company_id}_{session_id}.json"
    session = load_json(session_path, {"id": session_id, "company_id": company_id, "type": "interview", "messages": [], "created_at": datetime.now().isoformat()})

    agent = get_agent("financial_interviewer")
    if not agent:
        raise HTTPException(500, "Agente entrevistador no configurado")

    system_prompt = get_prompt("financial_interviewer")
    if agent.get("system_prompt"):
        system_prompt = agent["system_prompt"]

    messages = [{"role": "system", "content": system_prompt}]
    for msg in session.get("messages", []):
        messages.append({"role": msg["role"], "content": msg["content"]})
    messages.append({"role": "user", "content": data.message})

    response = call_ollama_chat(
        model=agent.get("model", "llama3.2:3b"),
        messages=messages,
        temperature=agent.get("temperature", 0.7)
    )

    session["messages"].append({"role": "user", "content": data.message, "timestamp": datetime.now().isoformat()})
    session["messages"].append({"role": "assistant", "content": response, "timestamp": datetime.now().isoformat()})
    save_json(session_path, session)

    if company.get("status") == "pending":
        company["status"] = "interviewing"
        company["updated_at"] = datetime.now().isoformat()
        save_json(company_dir / "company.json", company)

    return {
        "response": response,
        "session_id": session_id,
        "company_status": company.get("status")
    }

# ---------------------------------------------------------------------------
# Endpoints: Generación de Flujo de Caja
# ---------------------------------------------------------------------------
@app.post("/api/companies/{company_id}/generate-cashflow")
async def generate_cashflow(company_id: str, background_tasks: BackgroundTasks):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    if not company_dir.exists():
        raise HTTPException(404, "Empresa no encontrada")

    _check_rate_limit(f"gen_{company_id}", max_requests=5, window=120)

    task_id = str(uuid.uuid4())[:8]
    _generation_status[task_id] = {"status": "generating", "progress": 0, "error": None, "step": "Iniciando..."}

    background_tasks.add_task(_generate_cashflow_task, company_id, task_id)
    return {"task_id": task_id, "status": "generating"}

def _generate_cashflow_task(company_id: str, task_id: str):
    try:
        company_dir = DATA_DIR / "companies" / company_id
        company = load_json(company_dir / "company.json")

        _generation_status[task_id]["step"] = "Recopilando información de entrevistas..."
        _generation_status[task_id]["progress"] = 10

        sessions_dir = DATA_DIR / "sessions"
        all_messages = []
        for f in sorted(sessions_dir.glob(f"{company_id}_*.json")):
            session = load_json(f)
            for msg in session.get("messages", []):
                all_messages.append(f"{msg['role'].upper()}: {msg['content']}")

        conversation_text = "\n".join(all_messages)
        if len(conversation_text) > 50000:
            conversation_text = conversation_text[:50000] + "\n[... conversación truncada por longitud ...]"

        _generation_status[task_id]["progress"] = 25
        _generation_status[task_id]["step"] = "Preparando análisis financiero..."

        agent = get_agent("cashflow_analyst")
        if not agent:
            _generation_status[task_id] = {"status": "error", "progress": 0, "error": "Agente analista no configurado", "step": "Error"}
            return

        system_prompt = get_prompt("cashflow_analyst")
        if agent.get("system_prompt"):
            system_prompt = agent["system_prompt"]

        user_message = f"""Basándote en la siguiente conversación con el dueño de la empresa "{company.get('name', 'Sin nombre')}" del sector "{company.get('sector', 'No especificado')}",
construye un flujo de caja mensual proyectado a 12 meses.

CONVERSACIÓN COMPLETA:
{conversation_text}

Genera el JSON del flujo de caja siguiendo exactamente el formato indicado en tus instrucciones."""

        _generation_status[task_id]["progress"] = 40
        _generation_status[task_id]["step"] = "La IA está analizando los datos financieros..."

        response = call_ollama(
            model=agent.get("model", "llama3.1:8b"),
            system_prompt=system_prompt,
            user_message=user_message,
            temperature=agent.get("temperature", 0.3),
            timeout=_get_timeout("analysis")
        )

        _generation_status[task_id]["progress"] = 75
        _generation_status[task_id]["step"] = "Procesando resultados..."

        cashflow_data = _extract_json_from_llm(response)
        if not cashflow_data:
            # Intentar extraer datos de la respuesta como texto estructurado
            logger.warning(f"No se pudo extraer JSON del LLM para {company_id}. Respuesta: {response[:200]}")
            _generation_status[task_id] = {"status": "error", "progress": 0, "error": "No se pudo generar el flujo de caja. La IA no devolvió datos en formato correcto. Intenta proporcionar más información en la entrevista.", "step": "Error"}
            return

        _generation_status[task_id]["progress"] = 85
        _generation_status[task_id]["step"] = "Normalizando y validando datos..."

        cashflow_data = normalize_cashflow(cashflow_data)
        
        # Validar que se generaron meses
        num_months = len(cashflow_data.get("months", []))
        if num_months == 0:
            logger.warning(f"Cashflow generado sin meses para {company_id}. JSON extraído: {json.dumps(cashflow_data, default=str)[:500]}")
            _generation_status[task_id] = {"status": "error", "progress": 0, "error": "La IA generó un flujo de caja vacío (0 meses). Esto puede ocurrir si no hay suficiente información. Vuelve a la entrevista y proporciona más detalles sobre ingresos y gastos mensuales.", "step": "Error"}
            return
        
        # Validar que los totales no sean todos 0 (indicaría datos corruptos)
        total_income = cashflow_data.get("summary", {}).get("total_income", 0)
        total_expenses = cashflow_data.get("summary", {}).get("total_expenses", 0)
        if total_income == 0 and total_expenses == 0 and num_months > 0:
            logger.warning(f"Cashflow con {num_months} meses pero totales en 0 para {company_id}")
            # No es un error fatal, pero loguear para debug

        cashflow_data["id"] = str(uuid.uuid4())[:8]
        cashflow_data["company_id"] = company_id
        cashflow_data["created_at"] = datetime.now().isoformat()
        save_json(company_dir / "cashflow.json", cashflow_data)
        logger.info(f"Flujo de caja generado: {num_months} meses, ingresos={total_income:,.0f}, gastos={total_expenses:,.0f} para empresa {company_id}")

        company["status"] = "complete"
        company["updated_at"] = datetime.now().isoformat()
        save_json(company_dir / "company.json", company)

        _generation_status[task_id] = {"status": "done", "progress": 100, "error": None, "step": "Completado", "months_generated": num_months}

    except Exception as e:
        logger.error(f"Error generando flujo de caja para {company_id}: {type(e).__name__}")
        _generation_status[task_id] = {"status": "error", "progress": 0, "error": str(e), "step": "Error"}

@app.get("/api/generation/{task_id}/progress")
async def get_generation_progress(task_id: str):
    task_id = re.sub(r"[^a-zA-Z0-9_-]", "", task_id)[:64]
    return _generation_status.get(task_id, {"status": "unknown", "progress": 0, "error": "Tarea no encontrada", "step": ""})

@app.get("/api/companies/{company_id}/cashflow")
async def get_cashflow(company_id: str):
    company_id = _sanitize_id(company_id)
    path = DATA_DIR / "companies" / company_id / "cashflow.json"
    if not path.exists():
        raise HTTPException(404, "Flujo de caja no encontrado. Genera uno primero.")
    data = load_json(path)
    data = normalize_cashflow(data)
    return data

# ---------------------------------------------------------------------------
# Endpoints: Edición manual de meses del flujo de caja
# ---------------------------------------------------------------------------
@app.put("/api/companies/{company_id}/cashflow/months/{month_index}")
async def update_month(company_id: str, month_index: int, month_data: MonthData):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "No hay flujo de caja para editar.")

    cashflow = load_json(cashflow_path)
    months = cashflow.get("months", [])

    if month_index < 0 or month_index >= len(months):
        raise HTTPException(400, f"Índice de mes inválido: {month_index}. Rango válido: 0-{len(months)-1}")

    existing = months[month_index]
    if month_data.month:
        existing["month"] = month_data.month
    if month_data.label:
        existing["label"] = month_data.label
    if month_data.income is not None:
        existing["income"] = month_data.income
    if month_data.expenses is not None:
        existing["expenses"] = month_data.expenses

    months[month_index] = existing
    cashflow["months"] = months
    cashflow = normalize_cashflow(cashflow)
    save_json(cashflow_path, cashflow)

    return cashflow

@app.post("/api/companies/{company_id}/cashflow/months")
async def add_month(company_id: str, month_data: MonthData):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "No hay flujo de caja. Genera uno primero.")

    cashflow = load_json(cashflow_path)
    months = cashflow.get("months", [])

    if len(months) >= MAX_MONTHS:
        raise HTTPException(400, f"No se pueden agregar más de {MAX_MONTHS} meses.")

    new_month = {
        "month": month_data.month or "",
        "label": month_data.label or "",
        "income": month_data.income or {"sales": 0, "other_income": 0, "total": 0},
        "expenses": month_data.expenses or {"variable_costs": 0, "fixed_costs": 0, "variable_expenses": 0, "debt_payments": 0, "taxes": 0, "investments": 0, "total": 0}
    }
    months.append(new_month)
    cashflow["months"] = months
    cashflow = normalize_cashflow(cashflow)
    save_json(cashflow_path, cashflow)

    return cashflow

@app.delete("/api/companies/{company_id}/cashflow/months/{month_index}")
async def delete_month(company_id: str, month_index: int):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "No hay flujo de caja.")

    cashflow = load_json(cashflow_path)
    months = cashflow.get("months", [])

    if month_index < 0 or month_index >= len(months):
        raise HTTPException(400, f"Índice de mes inválido: {month_index}")

    months.pop(month_index)
    cashflow["months"] = months
    cashflow = normalize_cashflow(cashflow)
    save_json(cashflow_path, cashflow)

    return cashflow

# ---------------------------------------------------------------------------
# Endpoints: Simulación de Escenarios (OPTIMIZADA)
# ---------------------------------------------------------------------------
@app.post("/api/companies/{company_id}/simulate")
async def simulate_scenario(company_id: str, data: ScenarioRequest, background_tasks: BackgroundTasks):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "No hay flujo de caja base para simular.")

    _check_rate_limit(f"sim_{company_id}", max_requests=10, window=120)

    task_id = str(uuid.uuid4())[:8]
    _generation_status[task_id] = {
        "status": "generating",
        "progress": 0,
        "error": None,
        "step": "Iniciando simulación...",
        "mode": "local" if data.params else "ai"
    }

    if data.params:
        # Simulación local rápida (sin LLM)
        background_tasks.add_task(_simulate_local_task, company_id, data.instruction, data.params, task_id)
    else:
        # Simulación con IA (para instrucciones complejas en lenguaje natural)
        background_tasks.add_task(_simulate_ai_task, company_id, data.instruction, task_id)

    return {"task_id": task_id, "status": "generating"}


def _simulate_local_task(company_id: str, instruction: str, params: dict, task_id: str):
    """Simulación local rápida sin LLM."""
    try:
        company_dir = DATA_DIR / "companies" / company_id
        cashflow = load_json(company_dir / "cashflow.json")

        _generation_status[task_id]["progress"] = 10
        _generation_status[task_id]["step"] = "Cargando flujo de caja base..."

        # Ejecutar simulación matemática local
        scenario_data = simulate_local(cashflow, params, task_id)

        # Guardar escenario
        scenario_id = str(uuid.uuid4())[:8]
        scenario_data["id"] = scenario_id
        scenario_data["company_id"] = company_id
        scenario_data["instruction"] = instruction
        scenario_data["simulation_mode"] = "local"
        scenario_data["created_at"] = datetime.now().isoformat()

        scenarios_dir = DATA_DIR / "scenarios"
        scenarios_dir.mkdir(parents=True, exist_ok=True)
        save_json(scenarios_dir / f"{company_id}_{scenario_id}.json", scenario_data)

        _generation_status[task_id] = {
            "status": "done",
            "progress": 100,
            "error": None,
            "scenario_id": scenario_id,
            "step": "Simulación completada",
            "mode": "local"
        }
        logger.info(f"Simulación local completada para empresa {company_id}")

        # Limpiar checkpoint
        checkpoint_path = DATA_DIR / "checkpoints" / f"{task_id}.json"
        if checkpoint_path.exists():
            checkpoint_path.unlink()

    except Exception as e:
        logger.error(f"Error en simulación local para {company_id}: {type(e).__name__}: {e}")
        _generation_status[task_id] = {"status": "error", "progress": 0, "error": str(e), "step": "Error"}


def _simulate_ai_task(company_id: str, instruction: str, task_id: str):
    """Simulación con IA para instrucciones complejas en lenguaje natural."""
    try:
        company_dir = DATA_DIR / "companies" / company_id
        cashflow = load_json(company_dir / "cashflow.json")

        _generation_status[task_id]["progress"] = 15
        _generation_status[task_id]["step"] = "Preparando datos para la IA..."

        agent = get_agent("scenario_simulator")
        if not agent:
            _generation_status[task_id] = {"status": "error", "progress": 0, "error": "Agente simulador no configurado", "step": "Error"}
            return

        system_prompt = get_prompt("scenario_simulator")
        if agent.get("system_prompt"):
            system_prompt = agent["system_prompt"]

        # OPTIMIZACIÓN: Enviar solo un resumen compacto del cashflow, no el JSON completo
        months_compact = []
        for m in cashflow.get("months", []):
            months_compact.append({
                "month": m.get("month", ""),
                "label": m.get("label", ""),
                "income": m.get("income", {}),
                "expenses": m.get("expenses", {}),
                "net_flow": m.get("net_flow", 0),
                "cumulative_balance": m.get("cumulative_balance", 0)
            })

        cf_compact = {
            "months": months_compact,
            "summary": cashflow.get("summary", {})
        }

        user_message = f"""FLUJO DE CAJA ACTUAL:
{json.dumps(cf_compact, ensure_ascii=False)}

INSTRUCCIÓN DEL USUARIO:
{instruction}

Aplica los cambios solicitados y devuelve el flujo de caja completo actualizado en formato JSON."""

        _generation_status[task_id]["progress"] = 30
        _generation_status[task_id]["step"] = "La IA está calculando el escenario..."

        # Guardar checkpoint antes de llamar al LLM
        _save_simulation_checkpoint(task_id, cashflow, 0, len(cashflow.get("months", [])))

        response = call_ollama(
            model=agent.get("model", "llama3.1:8b"),
            system_prompt=system_prompt,
            user_message=user_message,
            temperature=agent.get("temperature", 0.4),
            timeout=_get_timeout("simulation")
        )

        _generation_status[task_id]["progress"] = 75
        _generation_status[task_id]["step"] = "Procesando respuesta de la IA..."

        scenario_data = _extract_json_from_llm(response)
        if not scenario_data:
            _generation_status[task_id] = {"status": "error", "progress": 0, "error": "La IA no pudo generar una simulación válida. Intenta con una instrucción más específica.", "step": "Error"}
            return

        _generation_status[task_id]["progress"] = 85
        _generation_status[task_id]["step"] = "Normalizando resultados..."

        # Normalizar el resultado
        scenario_data = normalize_cashflow(scenario_data)

        # Generar alertas si no las tiene
        if "alerts" not in scenario_data:
            alerts = []
            for i, m in enumerate(scenario_data.get("months", [])):
                if m.get("cumulative_balance", 0) < 0:
                    alerts.append({
                        "month": m.get("label", m.get("month", f"Mes {i+1}")),
                        "type": "danger",
                        "message": f"Saldo acumulado negativo: ${m['cumulative_balance']:,.0f}"
                    })
            scenario_data["alerts"] = alerts

        # Guardar escenario
        scenario_id = str(uuid.uuid4())[:8]
        scenario_data["id"] = scenario_id
        scenario_data["company_id"] = company_id
        scenario_data["instruction"] = instruction
        scenario_data["simulation_mode"] = "ai"
        scenario_data["created_at"] = datetime.now().isoformat()

        scenarios_dir = DATA_DIR / "scenarios"
        scenarios_dir.mkdir(parents=True, exist_ok=True)
        save_json(scenarios_dir / f"{company_id}_{scenario_id}.json", scenario_data)

        _generation_status[task_id] = {
            "status": "done",
            "progress": 100,
            "error": None,
            "scenario_id": scenario_id,
            "step": "Simulación completada",
            "mode": "ai"
        }
        logger.info(f"Simulación AI completada para empresa {company_id}")

        # Limpiar checkpoint
        checkpoint_path = DATA_DIR / "checkpoints" / f"{task_id}.json"
        if checkpoint_path.exists():
            checkpoint_path.unlink()

    except Exception as e:
        logger.error(f"Error en simulación AI para {company_id}: {type(e).__name__}: {e}")
        _generation_status[task_id] = {"status": "error", "progress": 0, "error": str(e), "step": "Error"}


@app.get("/api/companies/{company_id}/scenarios")
async def list_scenarios(company_id: str):
    company_id = _sanitize_id(company_id)
    scenarios_dir = DATA_DIR / "scenarios"
    scenarios_dir.mkdir(parents=True, exist_ok=True)
    scenarios = []
    for f in sorted(scenarios_dir.glob(f"{company_id}_*.json")):
        scenarios.append(load_json(f))
    return {"scenarios": scenarios}

@app.get("/api/scenarios/{scenario_id}")
async def get_scenario(scenario_id: str):
    scenario_id = _sanitize_id(scenario_id)
    scenarios_dir = DATA_DIR / "scenarios"
    for f in scenarios_dir.glob(f"*_{scenario_id}.json"):
        return load_json(f)
    raise HTTPException(404, "Escenario no encontrado")

@app.delete("/api/scenarios/{scenario_id}")
async def delete_scenario(scenario_id: str):
    scenario_id = _sanitize_id(scenario_id)
    scenarios_dir = DATA_DIR / "scenarios"
    for f in scenarios_dir.glob(f"*_{scenario_id}.json"):
        f.unlink()
        return {"status": "deleted"}
    raise HTTPException(404, "Escenario no encontrado")

# ---------------------------------------------------------------------------
# Endpoints: Chat de Simulación (conversacional)
# ---------------------------------------------------------------------------
@app.post("/api/chat/simulate")
async def chat_simulate(data: ChatMessage):
    company_id = _sanitize_id(data.company_id)
    company_dir = DATA_DIR / "companies" / company_id
    if not company_dir.exists():
        raise HTTPException(404, "Empresa no encontrada")

    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "Primero genera un flujo de caja base.")

    _check_rate_limit(f"simchat_{company_id}")

    cashflow = load_json(cashflow_path)

    session_id = data.session_id or f"sim_{str(uuid.uuid4())[:8]}"
    session_id = re.sub(r"[^a-zA-Z0-9_-]", "", session_id)[:64]
    session_path = DATA_DIR / "sessions" / f"{company_id}_{session_id}.json"
    session = load_json(session_path, {"id": session_id, "company_id": company_id, "type": "simulation", "messages": [], "created_at": datetime.now().isoformat()})

    agent = get_agent("scenario_simulator")
    if not agent:
        raise HTTPException(500, "Agente simulador no configurado")

    system_prompt = get_prompt("scenario_simulator")

    cf_summary = f"La empresa tiene un flujo de caja proyectado a {len(cashflow.get('months', []))} meses. "
    summary = cashflow.get("summary", {})
    cf_summary += f"Ingresos totales: ${summary.get('total_income', 0):,.0f}, Gastos totales: ${summary.get('total_expenses', 0):,.0f}, Flujo neto: ${summary.get('net_cashflow', 0):,.0f}."

    messages = [{"role": "system", "content": system_prompt + "\n\nCONTEXTO: " + cf_summary}]
    for msg in session.get("messages", []):
        messages.append({"role": msg["role"], "content": msg["content"]})
    messages.append({"role": "user", "content": data.message})

    response = call_ollama_chat(
        model=agent.get("model", "llama3.1:8b"),
        messages=messages,
        temperature=agent.get("temperature", 0.4),
        timeout=_get_timeout("simulation")
    )

    session["messages"].append({"role": "user", "content": data.message, "timestamp": datetime.now().isoformat()})
    session["messages"].append({"role": "assistant", "content": response, "timestamp": datetime.now().isoformat()})
    save_json(session_path, session)

    scenario_data = _extract_json_from_llm(response)
    has_scenario = scenario_data is not None and "months" in (scenario_data or {})

    return {
        "response": response,
        "session_id": session_id,
        "has_scenario": has_scenario,
        "scenario_data": scenario_data if has_scenario else None
    }

# ---------------------------------------------------------------------------
# Endpoints: Exportación (con nombres de archivo sanitizados)
# ---------------------------------------------------------------------------
@app.get("/api/companies/{company_id}/export/excel")
async def export_excel(company_id: str):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "No hay flujo de caja para exportar.")

    cashflow = load_json(cashflow_path)
    company = load_json(company_dir / "company.json")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado.")

    wb = Workbook()

    # --- Hoja 1: Flujo de Caja ---
    ws = wb.active
    ws.title = "Flujo de Caja"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D3DA6", end_color="0D3DA6", fill_type="solid")
    money_format = '#,##0'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Flujo de Caja — {company.get('name', 'Empresa')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="0D3DA6")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")

    headers = ["Mes", "Ingresos", "Costos Variables", "Costos Fijos", "Gastos Variables", "Deudas", "Impuestos", "Flujo Neto", "Saldo Acumulado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, month in enumerate(cashflow.get("months", []), 5):
        expenses = month.get("expenses", {})
        income = month.get("income", {})
        ws.cell(row=i, column=1, value=month.get("label", month.get("month", ""))).border = thin_border
        ws.cell(row=i, column=2, value=income.get("total", 0)).number_format = money_format
        ws.cell(row=i, column=3, value=expenses.get("variable_costs", 0)).number_format = money_format
        ws.cell(row=i, column=4, value=expenses.get("fixed_costs", 0)).number_format = money_format
        ws.cell(row=i, column=5, value=expenses.get("variable_expenses", 0)).number_format = money_format
        ws.cell(row=i, column=6, value=expenses.get("debt_payments", 0)).number_format = money_format
        ws.cell(row=i, column=7, value=expenses.get("taxes", 0)).number_format = money_format
        ws.cell(row=i, column=8, value=month.get("net_flow", 0)).number_format = money_format
        ws.cell(row=i, column=9, value=month.get("cumulative_balance", 0)).number_format = money_format
        for col in range(1, 10):
            ws.cell(row=i, column=col).border = thin_border

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    # --- Hoja 2: Resumen ---
    ws2 = wb.create_sheet("Resumen")
    summary = cashflow.get("summary", {})
    ws2["A1"] = "Resumen Financiero"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="0D3DA6")

    summary_data = [
        ("Ingresos Totales", summary.get("total_income", 0)),
        ("Gastos Totales", summary.get("total_expenses", 0)),
        ("Flujo Neto", summary.get("net_cashflow", 0)),
        ("Promedio Mensual", summary.get("average_monthly_balance", 0)),
    ]
    for i, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value).number_format = money_format

    alerts = cashflow.get("alerts", [])
    if alerts:
        ws2.cell(row=8, column=1, value="Alertas").font = Font(bold=True, size=12, color="DC2626")
        for i, alert in enumerate(alerts, 9):
            ws2.cell(row=i, column=1, value=alert.get("month", ""))
            ws2.cell(row=i, column=2, value=alert.get("message", ""))

    recs = cashflow.get("recommendations", [])
    if recs:
        row = 9 + len(alerts) + 1
        ws2.cell(row=row, column=1, value="Recomendaciones").font = Font(bold=True, size=12, color="3DAE2B")
        for i, rec in enumerate(recs, row + 1):
            ws2.cell(row=i, column=1, value=f"• {rec}")

    exports_dir = DATA_DIR / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    safe_name = _sanitize_filename(company.get("name", "empresa"))
    filename = f"flujo_caja_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = exports_dir / filename
    wb.save(str(filepath))

    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/companies/{company_id}/export/csv")
async def export_csv(company_id: str):
    company_id = _sanitize_id(company_id)
    company_dir = DATA_DIR / "companies" / company_id
    cashflow_path = company_dir / "cashflow.json"
    if not cashflow_path.exists():
        raise HTTPException(404, "No hay flujo de caja para exportar.")

    cashflow = load_json(cashflow_path)
    company = load_json(company_dir / "company.json")

    import csv

    exports_dir = DATA_DIR / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    safe_name = _sanitize_filename(company.get("name", "empresa"))
    filename = f"flujo_caja_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = exports_dir / filename

    with open(str(filepath), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Mes", "Ingresos", "Costos Variables", "Costos Fijos", "Gastos Variables", "Deudas", "Impuestos", "Flujo Neto", "Saldo Acumulado"])
        for month in cashflow.get("months", []):
            expenses = month.get("expenses", {})
            income = month.get("income", {})
            writer.writerow([
                month.get("label", month.get("month", "")),
                income.get("total", 0),
                expenses.get("variable_costs", 0),
                expenses.get("fixed_costs", 0),
                expenses.get("variable_expenses", 0),
                expenses.get("debt_payments", 0),
                expenses.get("taxes", 0),
                month.get("net_flow", 0),
                month.get("cumulative_balance", 0)
            ])

    return FileResponse(path=str(filepath), filename=filename, media_type="text/csv")


# Exportar escenario específico
@app.get("/api/scenarios/{scenario_id}/export/csv")
async def export_scenario_csv(scenario_id: str):
    scenario_id = _sanitize_id(scenario_id)
    scenarios_dir = DATA_DIR / "scenarios"
    scenario_data = None
    for f in scenarios_dir.glob(f"*_{scenario_id}.json"):
        scenario_data = load_json(f)
        break
    if not scenario_data:
        raise HTTPException(404, "Escenario no encontrado")

    import csv

    exports_dir = DATA_DIR / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    scenario_name = _sanitize_filename(scenario_data.get("scenario_name", "escenario"))
    filename = f"escenario_{scenario_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = exports_dir / filename

    with open(str(filepath), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Mes", "Ingresos", "Costos Variables", "Costos Fijos", "Gastos Variables", "Deudas", "Impuestos", "Flujo Neto", "Saldo Acumulado"])
        for month in scenario_data.get("months", []):
            expenses = month.get("expenses", {})
            income = month.get("income", {})
            writer.writerow([
                month.get("label", month.get("month", "")),
                income.get("total", 0),
                expenses.get("variable_costs", 0),
                expenses.get("fixed_costs", 0),
                expenses.get("variable_expenses", 0),
                expenses.get("debt_payments", 0),
                expenses.get("taxes", 0),
                month.get("net_flow", 0),
                month.get("cumulative_balance", 0)
            ])

    return FileResponse(path=str(filepath), filename=filename, media_type="text/csv")


@app.get("/api/scenarios/{scenario_id}/export/excel")
async def export_scenario_excel(scenario_id: str):
    scenario_id = _sanitize_id(scenario_id)
    scenarios_dir = DATA_DIR / "scenarios"
    scenario_data = None
    for f in scenarios_dir.glob(f"*_{scenario_id}.json"):
        scenario_data = load_json(f)
        break
    if not scenario_data:
        raise HTTPException(404, "Escenario no encontrado")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Escenario"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D3DA6", end_color="0D3DA6", fill_type="solid")
    money_format = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = scenario_data.get("scenario_name", "Escenario Simulado")
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="0D3DA6")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado: {scenario_data.get('created_at', datetime.now().isoformat())[:10]}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")

    headers = ["Mes", "Ingresos", "Costos Variables", "Costos Fijos", "Gastos Variables", "Deudas", "Impuestos", "Flujo Neto", "Saldo Acumulado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, month in enumerate(scenario_data.get("months", []), 5):
        expenses = month.get("expenses", {})
        income = month.get("income", {})
        ws.cell(row=i, column=1, value=month.get("label", month.get("month", ""))).border = thin_border
        ws.cell(row=i, column=2, value=income.get("total", 0)).number_format = money_format
        ws.cell(row=i, column=3, value=expenses.get("variable_costs", 0)).number_format = money_format
        ws.cell(row=i, column=4, value=expenses.get("fixed_costs", 0)).number_format = money_format
        ws.cell(row=i, column=5, value=expenses.get("variable_expenses", 0)).number_format = money_format
        ws.cell(row=i, column=6, value=expenses.get("debt_payments", 0)).number_format = money_format
        ws.cell(row=i, column=7, value=expenses.get("taxes", 0)).number_format = money_format
        ws.cell(row=i, column=8, value=month.get("net_flow", 0)).number_format = money_format
        ws.cell(row=i, column=9, value=month.get("cumulative_balance", 0)).number_format = money_format
        for col in range(1, 10):
            ws.cell(row=i, column=col).border = thin_border

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    # Hoja de cambios aplicados
    ws2 = wb.create_sheet("Cambios")
    ws2["A1"] = "Cambios Aplicados"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="0D3DA6")
    for i, change in enumerate(scenario_data.get("changes_applied", []), 3):
        ws2.cell(row=i, column=1, value=f"• {change}")

    impact = scenario_data.get("impact_summary", {})
    if isinstance(impact, dict):
        ws2.cell(row=len(scenario_data.get("changes_applied", [])) + 5, column=1, value="Impacto").font = Font(bold=True)
        ws2.cell(row=len(scenario_data.get("changes_applied", [])) + 6, column=1, value=impact.get("description", ""))

    exports_dir = DATA_DIR / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    scenario_name = _sanitize_filename(scenario_data.get("scenario_name", "escenario"))
    filename = f"escenario_{scenario_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = exports_dir / filename
    wb.save(str(filepath))

    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------------------------------------------------------
# Endpoints: Agentes
# ---------------------------------------------------------------------------
@app.get("/api/agents")
async def list_agents():
    return get_agents()

@app.get("/api/agents/{agent_id}")
async def get_agent_endpoint(agent_id: str):
    agent = get_agent(agent_id)
    if not agent:
        raise HTTPException(404, "Agente no encontrado")
    return agent

@app.put("/api/agents/{agent_id}")
async def update_agent(agent_id: str, config: AgentConfigUpdate):
    agent_id = _sanitize_id(agent_id)
    agents_data = get_agents()
    for i, a in enumerate(agents_data.get("agents", [])):
        if a["id"] == agent_id:
            if config.model is not None:
                agents_data["agents"][i]["model"] = config.model
            if config.temperature is not None:
                agents_data["agents"][i]["temperature"] = config.temperature
            if config.system_prompt is not None:
                agents_data["agents"][i]["system_prompt"] = config.system_prompt
            agents_data["agents"][i]["updated_at"] = datetime.now().isoformat()
            save_json(DATA_DIR / "agents" / "agents.json", agents_data)

            model = config.model or agents_data["agents"][i].get("model")
            if model and not _is_model_available(model):
                _start_pull_background(model)
                return {**agents_data["agents"][i], "pull_status": {"status": "queued", "model": model}}
            return agents_data["agents"][i]
    raise HTTPException(404, "Agente no encontrado")

@app.get("/api/models/status")
async def models_status():
    return {"pull_status": _pull_status}

@app.get("/api/models/available")
async def available_models():
    try:
        r = http_requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        return {"models": [m["name"] for m in r.json().get("models", [])]}
    except Exception:
        return {"models": []}

# ---------------------------------------------------------------------------
# Endpoints: Sesiones
# ---------------------------------------------------------------------------
@app.get("/api/companies/{company_id}/sessions")
async def list_sessions(company_id: str):
    company_id = _sanitize_id(company_id)
    sessions_dir = DATA_DIR / "sessions"
    sessions_dir.mkdir(parents=True, exist_ok=True)
    sessions = []
    for f in sorted(sessions_dir.glob(f"{company_id}_*.json")):
        s = load_json(f)
        sessions.append({
            "id": s.get("id"),
            "type": s.get("type", "interview"),
            "message_count": len(s.get("messages", [])),
            "created_at": s.get("created_at")
        })
    return {"sessions": sessions}

@app.get("/api/companies/{company_id}/sessions/{session_id}")
async def get_session(company_id: str, session_id: str):
    company_id = _sanitize_id(company_id)
    session_id = re.sub(r"[^a-zA-Z0-9_-]", "", session_id)[:64]
    session_path = DATA_DIR / "sessions" / f"{company_id}_{session_id}.json"
    if not session_path.exists():
        raise HTTPException(404, "Sesión no encontrada")
    session = load_json(session_path)
    return {
        "id": session.get("id"),
        "type": session.get("type", "interview"),
        "messages": session.get("messages", []),
        "created_at": session.get("created_at")
    }

# ---------------------------------------------------------------------------
# Montar archivos estáticos y arrancar
# ---------------------------------------------------------------------------
if APP_DIR.exists():
    app.mount("/ui", StaticFiles(directory=str(APP_DIR), html=True), name="ui")
else:
    logger.warning(f"Directorio de UI no encontrado: {APP_DIR}")

@app.on_event("startup")
async def startup():
    for d in ["agents", "prompts/system", "sessions", "exports", "companies", "cashflows", "scenarios", "checkpoints"]:
        (DATA_DIR / d).mkdir(parents=True, exist_ok=True)
    agents_dst = DATA_DIR / "agents" / "agents.json"
    if not agents_dst.exists():
        agents_src = DEFAULTS_DIR / "agents.json"
        if agents_src.exists():
            shutil.copy2(str(agents_src), str(agents_dst))
    prompts_dst = DATA_DIR / "prompts" / "system"
    prompts_src = DEFAULTS_DIR / "prompts"
    if prompts_src.exists():
        for f in prompts_src.glob("*.md"):
            dst = prompts_dst / f.name
            if not dst.exists():
                shutil.copy2(str(f), str(dst))
    threading.Thread(target=ensure_ollama_running, daemon=True).start()
    logger.info(f"CCS Cashflow Assistant v0.3.0 iniciado en puerto {PORT} ({sys.platform})")

@app.get("/")
async def root():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/ui/index.html")

if __name__ == "__main__":
    import uvicorn

    if sys.platform == "win32":
        import asyncio
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    host = "127.0.0.1"
    print(f"http://{host}:{PORT}")
    uvicorn.run(app, host=host, port=PORT, log_level="info")
